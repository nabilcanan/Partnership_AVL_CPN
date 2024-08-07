import openpyxl
from openpyxl.utils import get_column_letter


def add_concatenated_column(ws):
    # Find the columns for 'CLEAN MANUFACTURER NAME' and 'CPN'
    clean_manufacturer_col = None
    cpn_col = None

    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == 'CLEAN MANUFACTURER NAME':
            clean_manufacturer_col = col[0].column_letter
        elif col[0].value == 'CPN':
            cpn_col = col[0].column_letter

    # Ensure both columns are found
    if not clean_manufacturer_col or not cpn_col:
        print("Columns not found")
        return

    # Find the next open column for 'Concatenated'
    next_open_col_index = ws.max_column + 1
    ws.cell(row=1, column=next_open_col_index).value = 'Concatenated'

    # Populate the 'Concatenated' column with the formula
    for row in range(2, ws.max_row + 1):
        formula = f'=CONCATENATE({clean_manufacturer_col}{row}, "-", {cpn_col}{row})'
        ws.cell(row=row, column=next_open_col_index).value = formula

    # Remove duplicates in the 'Concatenated' column
    seen = set()
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=next_open_col_index)
        value = cell.value
        if value in seen:
            cell.value = None
        else:
            seen.add(value)


def add_cpn_ct_column(ws):
    # Find the 'CPN' column
    cpn_col = None
    cpn_col_index = None

    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == 'CPN':
            cpn_col = col[0].column_letter
            cpn_col_index = col[0].column

    # Ensure the 'CPN' column is found
    if not cpn_col:
        print("CPN column not found")
        return

    # Insert a new column next to the 'CPN' column
    ws.insert_cols(cpn_col_index + 1)
    cpn_ct_col_letter = get_column_letter(cpn_col_index + 1)
    ws.cell(row=1, column=cpn_col_index + 1).value = 'CPN Ct'

    # Populate the 'CPN Ct' column with 'Y' or 'N'
    cpn_values = {}
    for row in range(2, ws.max_row + 1):
        cpn_value = ws.cell(row=row, column=cpn_col_index).value
        if cpn_value in cpn_values:
            cpn_values[cpn_value].append(row)
        else:
            cpn_values[cpn_value] = [row]

    for rows in cpn_values.values():
        value = 'Y' if len(rows) > 1 else 'N'
        for row in rows:
            ws.cell(row=row, column=cpn_col_index + 1).value = value


def process_workbook(file_path, save_path):
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Add concatenated column and CPN Ct column
    add_concatenated_column(ws)
    add_cpn_ct_column(ws)

    # Enable auto filter for the first row
    ws.auto_filter.ref = ws.dimensions

    # Save the workbook to the specified save_path
    wb.save(save_path)
